# DOCS:
# https://clickup.com/api/clickupreference/operation/Gettimeentryhistory/

# Antes de rodar esse código, acesse esse link e pegue o código:
# https://app.clickup.com/api?client_id=9BQNB7IFOMPHHLCGVVE3R2R36H59AOYK&redirect_uri=https://www.2solve.com/

import requests
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

team_id = "3075996"
header = [
    "SETOR", "TAREFA", "RESPONSÁVEL", "TEMPO ESTIMADO (h)", "CUSTO ESTIMADO",
    "TEMPO RASTREADO (h)", "CUSTO REAL", "STATUS", "% CONCLUÍDO",
]

custo_funcionario = {
    "Ana Luiza": 21.93,
    "Anderson Ferreira": 14.41,
    "Bernardo Zampirole Brandão": 29.39,
    "Bianca Aguiar": 20.37,
    "Dayane Erlacher": 37.50,
    "Fernando Bisi Vieira": 29.50,
    "Franco Louzada": 37.50,
    "Gabriel da Silva Biancardi": 21.79,
    "Gabriel Farias": 13.30,
    "GUSTAVO FRINHANI": 62.82,
    "Henrique Loss Lopes": 13.30,
    "Henrique Puppim": 35.90,
    "João Fernando Rangel Guimarães": 21.79,
    "João Pedro Souza Rocha": 33.45,
    "Mariana Gomes Calheiros": 21.55,
    "Mateus Ruy Soares Gaudio": 25.65,
    "Mauricio Calheiros": 66.82,
    "Melquisedeque Shaloon Bento da Silva Gomes": 50.37,
    "Menno": 62.27,
    "Michel Carvalho": 34.20,
    "Pedro Vieira Lopes": 19.65,
    "Phelipe Augusto": 36.83,
    "Rafael Antunes Costa": 27.19,
    "Rodrigo Merigueti": 33.51,
    "Waldelicio Junior": 62.27,
    "Willian Pacheco Silva": 33.46,
    "Ricardo Calheiros": 89.55,
    "Luc Dijkstra": 20.91
}

data_hoje = datetime.now().strftime("%Y-%m-%d")
space_exclude = ["90131034562", "90131103749", "90131210570", "90131669969", "90131064862", "90130523790", "90131679449", "90130182730"]
# id: 90131034562 nome: Marketing
# id: 90131103749 nome: Comercial
# id: 90131210570 nome: Gestão do Conhecimento
# id: 90131669969 nome: Processos e Qualidade
# id: 90131064862 nome: Gerente Gustavo
# id: 90130523790 nome: Gerente Maurício
# id: 90131679449 nome: Gerente Júnior
# id: 90130182730 nome: Gerente Dayane


folder_exclude = ["90132771240", "90132869998", "90132004295", "90132676002", "90131094720"]
# id: 90132771240 nome: Brain Storm
# id: 90132869998 nome: Consultoria GSMAS
# id: 90132004295 nome: Gestão de projetos
# id: 90132676002 nome: TI
# id: 90131094720 nome: Tarefas Gerais

# code = "EWOVWCI2APY52Z57XDP79O88R0JS96SY"

# reqUrl = "https://app.clickup.com/api/v2/oauth/token"

# headersList = {
#  "Accept":.*/*,
#  "User-Agent": Thunder Client (https://www.thunderclient.co.),
#  "Content-Type": pplication.jso"
# }

# payload = json.dumps({
#   "client_id":.9BQNB7IFOMPHHLCGVVE3R2R36H59AOYK,
#   "client_secret":.YBR6BQJEACHDHELWHT4OZ16H49LWVJ2B51SH1DREN03BWP7RVMK6O73C1CI6SBPR,
#   "code": code,
#   "redirect_uri": https://www.2solve.com"
# })

# response = requests.request("POST", reqUrl, data=payload,  headers=headersList)

# response = response.json()
# if "access_token" in response:
#   print(response["access_token"])
# else:
#   print(response)

response = {}
response["access_token"] = (
    "3138343_6fada438d889343a074632a34b7c61afcff6e35c1bc63ba13127367756e4d2b2"
)

headers = {"Authorization": f'Bearer {response["access_token"]}', 'Content-Type': 'application/json'}

# Funções Auxiliares
def obter_dados_api(url):
    response = requests.get(url, headers=headers)
    return response.json() if response.status_code == 200 else []

def calcular_custos(task_name, responsaveis, time_estimate, time_spent):
    if not responsaveis:
        # if time_estimate != 0 or time_spent != 0:
        #     print(f"Tarefa sem responsável - {task_name}")
        return 0, 0

    custo = sum(custo_funcionario.get(nome.strip(), 0) for nome in responsaveis.split(";"))
    return time_estimate * custo, time_spent * custo


def processar_tarefa(tarefa, aba, prefixo=""):
    responsaveis = ";".join(assignee["username"] for assignee in tarefa["assignees"])
    time_estimate = float(tarefa.get("time_estimate", 0) or 0) / 3600000
    time_spent = float(tarefa.get("time_spent", 0) or 0) / 3600000
    cost_estimate, cost_spent = calcular_custos(tarefa["name"], responsaveis, time_estimate, time_spent)
    status = tarefa["status"]["status"]
    percent_concluido = 1 if status.lower() in {"concluídas", "fechadas"} else 0

    aba.append([
        tarefa["list"]["name"], f"{prefixo} {tarefa['name']}", responsaveis, time_estimate,
        cost_estimate, time_spent, cost_spent, status, percent_concluido
    ])

    detail_task = obter_dados_api(f"https://api.clickup.com/api/v2/task/{task['id']}?custom_task_ids=true&include_subtasks=true&include_markdown_description=true&custom_fields=string")

    return detail_task.get("subtasks", [])

def processar_subtarefas(name_tarefa, tarefa, aba, prefixo=""):
    responsaveis = ";".join(assignee["username"] for assignee in tarefa["assignees"])
    time_estimate = float(tarefa.get("time_estimate", 0) or 0) / 3600000
    time_spent = float(tarefa.get("time_spent", 0) or 0) / 3600000
    cost_estimate, cost_spent = calcular_custos(tarefa["name"], responsaveis, time_estimate, time_spent)
    status = tarefa["status"]["status"]
    percent_concluido = 1 if status.lower() in {"concluídas", "fechadas"} else 0

    aba.append([
        name_tarefa, f"{prefixo} {tarefa['name']}", responsaveis, time_estimate,
        cost_estimate, time_spent, cost_spent, status, percent_concluido
    ])

    if prefixo.count(".") == 4:
        return []
    
    detail_task = obter_dados_api(f"https://api.clickup.com/api/v2/task/{tarefa['id']}?custom_task_ids=true&include_subtasks=true&include_markdown_description=true&custom_fields=string")
    
    return detail_task.get("subtasks", [])  


try:
    spaces = obter_dados_api(f"https://api.clickup.com/api/v2/team/{team_id}/space?archived=false")["spaces"]

    for space in spaces:
        if space["id"] in space_exclude:
            continue
        space_name = space["name"]

        folders = obter_dados_api(f"https://api.clickup.com/api/v2/space/{space['id']}/folder?archived=false")["folders"]

        for folder in folders:
            if folder["id"] in folder_exclude:
                continue
            # print(folder["id"])
            # print(f'{space["name"]}_{folder["name"]}')
            # continue
            caminho_arquivo = f"./Projetos/{space["name"]}_{folder["name"]}.xlsx"
            print(caminho_arquivo)
            workbook = load_workbook(caminho_arquivo)
            if "Tarefas Gerais" in workbook.sheetnames:
                aba = workbook["Tarefas Gerais"]
            else:
                raise ValueError("A aba 'Tarefas Gerais' não existe na planilha.")

            # Limpa a aba e insere cabeçalho
            aba.delete_rows(1, aba.max_row)
            aba.append(header)

            # Obtém listas e processa tarefas
            lists = obter_dados_api(f"https://api.clickup.com/api/v2/folder/{folder["id"]}/list?archived=false")["lists"]
            for lista in lists:
                tasks = obter_dados_api(f"https://api.clickup.com/api/v2/list/{lista['id']}/task?include_closed=true")["tasks"]
                for idx, task in enumerate(tasks, start=1):
                    subtasks = processar_tarefa(task, aba, f"{idx}")
                    for sub_idx, subtask in enumerate(subtasks, start=1):
                        subtask2 = processar_subtarefas(task["list"]["name"], subtask, aba, f"{idx}.{sub_idx}.")
                        for sub2_idx, subtask2 in enumerate(subtask2, start=1):
                            subtask3 = processar_subtarefas(task["list"]["name"], subtask2, aba, f"{idx}.{sub_idx}.{sub2_idx}.")
                            for sub3_idx, subtask3 in enumerate(subtask3, start=1):
                                processar_subtarefas(task["list"]["name"], subtask3, aba, f"{idx}.{sub_idx}.{sub2_idx}.{sub3_idx}.")

            workbook.save(caminho_arquivo)
            # break
        # break

except requests.RequestException as e:
    print(f"Erro de requisição: {e}")