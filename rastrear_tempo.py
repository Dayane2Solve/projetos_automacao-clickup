import requests
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

team_id = "3075996"
response = {}
response["access_token"] = (
    "3138343_6fada438d889343a074632a34b7c61afcff6e35c1bc63ba13127367756e4d2b2"
)

headers = {
    "Authorization": f'Bearer {response["access_token"]}',
    "Content-Type": "application/json",
}

ids_usuarios = [
    {"id": 82184428, "nome_completo": "Thiago Ribeiro Pompermayer", "carga": 8, "departamento": "Comercial", "nome": "Thiago"},
    {"id": 82174345, "nome_completo": "Gabriel Farias", "carga": 4, "departamento": "Hardware", "nome": "Gabriel Farias"},
    {"id": 82173689, "nome_completo": "Mateus Ruy Soares Gaudio", "carga": 8, "departamento": "Firmware", "nome": "Mateus"},
    {"id": 82164056, "nome_completo": "João Pedro Souza Rocha", "carga": 8, "departamento": "Software", "nome": "João Pedro"},
    {"id": 82155499, "nome_completo": "Pedro Vieira Lopes", "carga": 5, "departamento": "Software", "nome": "Pedro"},
    {"id": 82150464, "nome_completo": "Mariana Gomes Calheiros", "carga": 2.4, "departamento": "Qualidade", "nome": "Mariana"},
    {"id": 82145319, "nome_completo": "Henrique Loss Lopes", "carga": 4, "departamento": "Mecânica", "nome": "Henrique"},
    # {"id": 3138354, "nome_completo": "Waldelicio Junior", "carga": 8, "departamento": "Produção", "nome": "Waldelicio"},
    {"id": 82083217, "nome_completo": "João Fernando Rangel Guimarães", "carga": 4, "departamento": "Firmware", "nome": "João Fernando"},
    {"id": 82074328, "nome_completo": "Michel Carvalho", "carga": 8, "departamento": "Comercial", "nome": "Michel"},
    {"id": 89143754, "nome_completo": "Anderson Ferreira", "carga": 6, "departamento": "Marketing", "nome": "Anderson"},
    {"id": 60976772, "nome_completo": "Ana Luiza", "carga": 8, "departamento": "Marketing", "nome": "Ana"},
    # {"id": 3138012, "nome_completo": "Ricardo Calheiros", "carga": 8, "departamento": "Comercial", "nome": "Ricardo"},
    # {"id": 84164888, "nome_completo": "Bianca Aguiar", "carga": 8, "departamento": "Projetos", "nome": "Bianca"},
    {"id": 82042179, "nome_completo": "Gabriel da Silva Biancardi", "carga": 4, "departamento": "Firmware", "nome": "Gabriel Biancardi"},
    {"id": 84636582, "nome_completo": "Willian Pacheco Silva", "carga": 8, "departamento": "Software", "nome": "Willian"},
    {"id": 84636590, "nome_completo": "Bernardo Zampirole Brandão", "carga": 8, "departamento": "Software", "nome": "Bernardo"},
    {"id": 81917609, "nome_completo": "Rafael Antunes Costa", "carga": 8, "departamento": "Software", "nome": "Rafael"},
    # {"id": 3138355, "nome_completo": "GUSTAVO FRINHANI", "carga": 8, "departamento": "Comissionamento", "nome": "GUSTAVO"},
    {"id": 3138350, "nome_completo": "Melquisedeque Shaloon Bento da Silva Gomes", "carga": 8, "departamento": "Hardware", "nome": "Melquisedeque"},
    {"id": 60932018, "nome_completo": "Fernando Bisi Vieira", "carga": 8, "departamento": "Firmware", "nome": "Fernando"},
    {"id": 3269111, "nome_completo": "Henrique Puppim", "carga": 8, "departamento": "Mecânica", "nome": "Henrique"},
    {"id": 3248906, "nome_completo": "Rodrigo Merigueti", "carga": 8, "departamento": "Software", "nome": "Rodrigo"},
    # {"id": 3219862, "nome_completo": "Mauricio Calheiros", "carga": 8, "departamento": "P&D", "nome": "Mauricio"},
    {"id": 3164413, "nome_completo": "Phelipe Augusto", "carga": 8, "departamento": "Software", "nome": "Phelipe"},
    # {"id": 3138343, "nome_completo": "Dayane Erlacher", "carga": 8, "departamento": "Projetos", "nome": "Dayane"},
    # {"id": 3138065, "nome_completo": "Franco Louzada", "carga": 8, "departamento": "Software", "nome": "Franco"},
    {"id": 3137926, "nome_completo": "Menno", "carga": 8, "departamento": "Firmware", "nome": "Menno"},
]


# Funções Auxiliares
def obter_dados_api(url):
    response = requests.get(url, headers=headers)
    return response.json() if response.status_code == 200 else []

try:
    strdata = "01/11/2024"
    datainicio = datetime.strptime(strdata, "%d/%m/%Y")
    strdata = "30/11/2024"
    strdatafim = datetime.strptime(strdata, "%d/%m/%Y")
    dias_uteis = 20
    start_date = int(datainicio.timestamp() * 1000)
    end_date = int(strdatafim.timestamp() * 1000)
    caminho_arquivo = "../Planilhas/tempo_rastreado.xlsx"
    workbook = load_workbook(caminho_arquivo)

    if "Descritivo" in workbook.sheetnames:
        aba = workbook["Descritivo"]
    else:
        raise ValueError("A aba 'Descritivo' não existe na planilha.")
    
    aba.delete_rows(2, aba.max_row)


    for id_user in ids_usuarios:
        detaild_user = obter_dados_api(
            f"https://api.clickup.com/api/v2/team/{team_id}/time_entries?start_date={start_date}&end_date={end_date}&assignee={id_user["id"]}"
        )["data"]
        time_count = 0
        for user in detaild_user:
            time_count += int(user["duration"])
        time_count_total = dias_uteis * id_user["carga"]
        aba.append(
            [
                id_user["nome"],
                time_count / 3600000,
                (time_count / 3600000) / time_count_total,
                id_user["departamento"]
            ]
        )
    workbook.save(caminho_arquivo)
except requests.RequestException as e:
    print(f"Erro de requisição: {e}")