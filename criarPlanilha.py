# DOCS:
# https://clickup.com/api/clickupreference/operation/Gettimeentryhistory/

# Antes de rodar esse código, acesse esse link e pegue o código:
# https://app.clickup.com/api?client_id=9BQNB7IFOMPHHLCGVVE3R2R36H59AOYK&redirect_uri=https://www.2solve.com/

import requests
from openpyxl import load_workbook
from datetime import datetime

team_id = "3075996"
header = [
    "SETOR/SPRINT", "ID", "TAREFA", "RESPONSÁVEL", "TEMPO ESTIMADO INICIAL (h)", "CUSTO ESTIMADO INICIAL", "TEMPO ESTIMADO (h)", "CUSTO ESTIMADO",
    "TEMPO RASTREADO (h)", "CUSTO REAL", "STATUS", "% CONCLUÍDO", "DATA CRIADA",	"DATA ATUALIZADA",	"DATA CONCLUÍDA", "DATA FECHADA", "Horizonte de inovação"
]

custo_funcionario = {
    "Ana Luiza": 21.93,
    "Anderson Ferreira": 14.41,
    "Bernardo Zampirole Brandão": 29.39,
    "Bianca Aguiar": 20.37,
    "Dayane Erlacher": 37.50,
    "Fernando Bisi Vieira": 29.50,
    "Franco Louzada": 37.50,
    "Gabriel da Silva Biancardi": 21.79,
    "Gabriel Farias": 13.30,
    "GUSTAVO FRINHANI": 62.82,
    "Henrique Loss Lopes": 13.30,
    "Henrique Puppim": 35.90,
    "João Fernando Rangel Guimarães": 21.79,
    "João Pedro Souza Rocha": 33.45,
    "Mariana Gomes Calheiros": 21.55,
    "Mateus Ruy Soares Gaudio": 25.65,
    "Mauricio Calheiros": 66.82,
    "Melquisedeque Shaloon Bento da Silva Gomes": 50.37,
    "Menno": 62.27,
    "Michel Carvalho": 34.20,
    "Pedro Vieira Lopes": 19.65,
    "Phelipe Augusto": 36.83,
    "Rafael Antunes Costa": 27.19,
    "Rodrigo Merigueti": 33.51,
    "Waldelicio Junior": 62.27,
    "Willian Pacheco Silva": 33.46,
    "Ricardo Calheiros": 89.55,
    "Luc Dijkstra": 20.91
}

data_hoje = datetime.now().strftime("%Y-%m-%d")

response = {}
response["access_token"] = (
    "3138343_6fada438d889343a074632a34b7c61afcff6e35c1bc63ba13127367756e4d2b2"
)

headers = {"Authorization": f'Bearer {response["access_token"]}', 'Content-Type': 'application/json'}

# Funções Auxiliares
def obter_dados_api(url):
    response = requests.get(url, headers=headers)
    return response.json() if response.status_code == 200 else []

def calcular_custos(task_name, responsaveis, time_estimate, time_spent):
    if not responsaveis:
        # if time_estimate != 0 or time_spent != 0:
        #     print(f"Tarefa sem responsável - {task_name}")
        return 0, 0

    custo = sum(custo_funcionario.get(nome.strip(), 0) for nome in responsaveis.split(";"))
    return time_estimate * custo, time_spent * custo


def processar_tarefa(nome_tarefa, tarefa, aba, prefixo="", nivel=1, max_nivel=4):
    if nivel > max_nivel:
        return []

    responsaveis = ";".join(assignee["username"] for assignee in tarefa["assignees"])
    time_estimate = float(tarefa.get("time_estimate", 0) or 0) / 3600000
    time_spent = float(tarefa.get("time_spent", 0) or 0) / 3600000
    cost_estimate, cost_spent = calcular_custos(tarefa["name"], responsaveis, time_estimate, time_spent)
    status = tarefa["status"]["status"]
    percent_concluido = 1 if status.lower() in {"concluídas", "fechadas"} else 0
    date_created = (datetime.fromtimestamp(int(tarefa.get("date_created")) / 1000)).strftime('%d/%m/%Y %H:%M:%S') if tarefa.get("date_created") else ''
    date_updated = (datetime.fromtimestamp(int(tarefa.get("date_updated")) / 1000)).strftime('%d/%m/%Y %H:%M:%S') if tarefa.get("date_updated") else ''
    date_closed = (datetime.fromtimestamp(int(tarefa.get("date_closed")) / 1000)).strftime('%d/%m/%Y %H:%M:%S') if tarefa.get("date_closed") else ''
    date_done = (datetime.fromtimestamp(int(tarefa.get("date_done")) / 1000)).strftime('%d/%m/%Y %H:%M:%S') if tarefa.get("date_done") else ''
    h = ''
    for horizonte in tarefa.get("custom_fields", []):
        if horizonte["id"] == "7c2b82c3-af45-4074-8633-ad9ee008c6a2" and "value" in horizonte:
            if horizonte["value"] == 0:
                h = "H1"
            elif horizonte["value"] == 1:
                h = "H2"
            elif horizonte["value"] == 2:
                h = "H3"

    aba.append([nome_tarefa, tarefa['id'], f"{prefixo} {tarefa['name']}", responsaveis, time_estimate, cost_estimate, time_estimate, cost_estimate, time_spent, cost_spent, status, percent_concluido, date_created, date_updated, date_closed, date_done, h])

    detail_task = obter_dados_api(f"https://api.clickup.com/api/v2/task/{tarefa['id']}?custom_task_ids=true&include_subtasks=true&include_markdown_description=true")

    if isinstance(detail_task, dict):
        subtasks = detail_task.get("subtasks", [])
    else:
        subtasks = []
    
    for sub_idx, subtask in enumerate(subtasks, start=1):
        processar_tarefa(nome_tarefa, subtask, aba, f"{prefixo}.{sub_idx}", nivel=nivel + 1, max_nivel=max_nivel)

    return subtasks

# Verificar se a planilha está vazia
def is_sheet_empty(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None:  # Se encontrar uma célula preenchida
                return False
    return True

try:
    folder_id = 90132453386
    caminho_arquivo = f"./Projetos/Overview_IoTTruck White Martins.xlsx"
    workbook = load_workbook(caminho_arquivo)
    if "Tarefas Gerais" in workbook.sheetnames:
        aba = workbook["Tarefas Gerais"]
    else:
        raise ValueError("A aba 'Tarefas Gerais' não existe na planilha.")

    aba.delete_rows(1, aba.max_row)
    aba.append(header)
    lists = obter_dados_api(f"https://api.clickup.com/api/v2/folder/{folder_id}/list?archived=false")["lists"]
    for lista in lists:
        tasks = obter_dados_api(f"https://api.clickup.com/api/v2/list/{lista['id']}/task?include_closed=true")["tasks"]
        for idx, task in enumerate(tasks, start=1):
            subtasks = processar_tarefa(task["list"]["name"], task, aba, f"{idx}.")
    
    
    workbook.save(caminho_arquivo)


except requests.RequestException as e:
    print(f"Erro de requisição: {e}")