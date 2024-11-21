import requests
import json
from openpyxl import load_workbook
from datetime import datetime

# Configurações
team_id = "3075996"
header = [
    "SETOR", "TAREFA", "RESPONSÁVEL", "TEMPO ESTIMADO (h)", "CUSTO ESTIMADO",
    "TEMPO RASTREADO (h)", "CUSTO REAL", "STATUS", "% CONCLUÍDO",
]
custo_funcionario = {
    "Ana Luiza": 21.93, "Anderson Ferreira": 14.41, "Bernardo Zampirole Brandão": 29.39,
    "Bianca Aguiar": 20.37, "Dayane Erlacher": 47.64, "Fernando Bisi Vieira": 29.50,
    "Franco Louzada": 49.75, "Gabriel da Silva Biancardi": 21.79, "Gabriel Farias": 21.79,
    "GUSTAVO FRINHANI": 76.45, "Henrique Loss Lopes": 21.79, "Henrique Puppim": 32.04,
    "João Fernando Rangel Guimarães": 24.39, "João Pedro Souza Rocha": 29.59,
    "Mariana Gomes Calheiros": 22.23, "Mateus Ruy Soares Gaudio": 29.39,
    "Mauricio Calheiros": 83.57, "Melquisedeque Shaloon Bento da Silva Gomes": 47.64,
    "Menno": 75.68, "Michel Carvalho": 34.09, "Pedro Vieira Lopes": 21.23,
    "Phelipe Augusto": 36.83, "Rafael Antunes Costa": 29.39, "Rodrigo Merigueti": 33.51,
    "Waldelicio Junior": 93.43, "Willian Pacheco Silva": 33.46, "Ricardo Calheiros": 109.10,
    "Luc Dijkstra": 20.91,
}
data_hoje = datetime.now().strftime("%Y-%m-%d")
token = "3138343_6fada438d889343a074632a34b7c61afcff6e35c1bc63ba13127367756e4d2b2"
headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

# Funções Auxiliares
def obter_dados_api(url):
    response = requests.get(url, headers=headers)
    return response.json() if response.status_code == 200 else None

def calcular_custos(task_name, responsaveis, time_estimate, time_spent):
    if not responsaveis:
        print(f"Tarefa sem responsável - {task_name}")
        return 0, 0

    custo = sum(custo_funcionario.get(nome.strip(), 0) for nome in responsaveis.split(";"))
    return time_estimate * custo, time_spent * custo


def processar_tarefa(tarefa, aba, prefixo=""):
    responsaveis = ";".join(assignee["username"] for assignee in tarefa["assignees"])
    time_estimate = float(tarefa.get("time_estimate", 0) or 0) / 3600000
    time_spent = float(tarefa.get("time_spent", 0) or 0) / 3600000
    cost_estimate, cost_spent = calcular_custos(tarefa["name"], responsaveis, time_estimate, time_spent)
    status = tarefa["status"]["status"]
    percent_concluido = 1 if task["status"]["status"].lower() in {"concluídas", "fechadas"} else 0

    aba.append([
        tarefa["list"]["name"], f"{prefixo} {tarefa['name']}", responsaveis, time_estimate,
        cost_estimate, time_spent, cost_spent, status, percent_concluido
    ])

    detail_task = obter_dados_api(f"https://api.clickup.com/api/v2/task/{task['id']}?custom_task_ids=true&include_subtasks=true&include_markdown_description=true&custom_fields=string")

    return detail_task.get("subtasks", [])

def processar_subtarefas(name_tarefa, tarefa, aba, prefixo=""):
    responsaveis = ";".join(assignee["username"] for assignee in tarefa["assignees"])
    time_estimate = float(tarefa.get("time_estimate", 0) or 0) / 3600000
    time_spent = float(tarefa.get("time_spent", 0) or 0) / 3600000
    cost_estimate, cost_spent = calcular_custos(tarefa["name"], responsaveis, time_estimate, time_spent)
    status = tarefa["status"]["status"]

    aba.append([
        name_tarefa, f"{prefixo} {tarefa['name']}", responsaveis, time_estimate,
        cost_estimate, time_spent, cost_spent, status, ""
    ])

    if prefixo.count(".") == 4:
        return []
    
    detail_task = obter_dados_api(f"https://api.clickup.com/api/v2/task/{tarefa['id']}?custom_task_ids=true&include_subtasks=true&include_markdown_description=true&custom_fields=string")
    if detail_task:
        return detail_task.get("subtasks", [])
    
    return []

# Execução Principal
try:
    id_folder = "90131110949"
    caminho_arquivo = "./Projetos/Gerente Menno_IoTTruck.xlsx"
    workbook = load_workbook(caminho_arquivo)
    if "Tarefas Gerais" in workbook.sheetnames:
        aba = workbook["Tarefas Gerais"]
    else:
        raise ValueError("A aba 'Tarefas Gerais' não existe na planilha.")

    # Limpa a aba e insere cabeçalho
    aba.delete_rows(1, aba.max_row)
    aba.append(header)

    # Obtém listas e processa tarefas
    lists = obter_dados_api(f"https://api.clickup.com/api/v2/folder/{id_folder}/list?archived=false")["lists"]
    for lista in lists:
        tasks = obter_dados_api(f"https://api.clickup.com/api/v2/list/{lista['id']}/task?include_closed=true")["tasks"]
        for idx, task in enumerate(tasks, start=1):
            subtasks = processar_tarefa(task, aba, f"{idx}")
            for sub_idx, subtask in enumerate(subtasks, start=1):
                subtask2 = processar_subtarefas(task["list"]["name"], subtask, aba, f"{idx}.{sub_idx}.")
                for sub2_idx, subtask2 in enumerate(subtask2, start=1):
                    subtask3 = processar_subtarefas(task["list"]["name"], subtask2, aba, f"{idx}.{sub_idx}.{sub2_idx}.")
                    for sub3_idx, subtask3 in enumerate(subtask3, start=1):
                        processar_subtarefas(task["list"]["name"], subtask3, aba, f"{idx}.{sub_idx}.{sub2_idx}.{sub3_idx}.")

    # Salva o arquivo
    workbook.save(caminho_arquivo)

except requests.RequestException as e:
    print(f"Erro de requisição: {e}")
